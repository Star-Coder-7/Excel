from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
# wb = Workbook()

# load an existing spreadsheet
wb = load_workbook('hello.xlsx')

# create an active worksheet
ws = wb.active

# Set variables
# name = ws['A3'].value
# color = ws['B3'].value

# print something from our spreadsheet
# print(f"{name}: {color}")

# Grab a whole column
columnA = ws['7']

for cell in columnA:
    print(cell.value)